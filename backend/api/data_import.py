# -*- coding: utf-8 -*-
"""
数据导入模块 - 支持从Excel导入并合并8个模块的数据
使用 openpyxl 直接处理 Excel 文件
"""
from flask import request, jsonify
from datetime import datetime
from sqlalchemy import and_
from . import api_bp
from models import db, Company, ProfitRate, NonRecurring, RoeNetAsset, PeValuation
from models import ShareholderStructure, ShareholderCount, RdExpense, RdStaff


MODULE_SCHEMA = {
    'profit_rate': {
        'required_fields': ['代码', '个股名称', '年份'],
        'optional_fields': ['销售毛利率(%)', '销售净利率(%)'],
        'model': ProfitRate,
        'key_fields': ['company_id', 'year'],
        'field_mapping': {
            '销售毛利率(%)': 'gross_profit_margin',
            '销售净利率(%)': 'net_profit_margin'
        },
        'sheet_names': ['利润率', '毛利率']
    },
    'non_recurring': {
        'required_fields': ['代码', '个股名称', '年份'],
        'optional_fields': ['扣非净利润(亿元)', '扣非增长率'],
        'model': NonRecurring,
        'key_fields': ['company_id', 'year'],
        'field_mapping': {
            '扣非净利润(亿元)': 'non_recurring_profit',
            '扣非增长率': 'non_recurring_growth'
        },
        'sheet_names': ['扣非净利润', '扣非']
    },
    'roe_net_asset': {
        'required_fields': ['代码', '个股名称', '年份'],
        'optional_fields': ['ROE(%)', '每股净资产(元)'],
        'model': RoeNetAsset,
        'key_fields': ['company_id', 'year'],
        'field_mapping': {
            'ROE(%)': 'roe',
            '每股净资产(元)': 'net_asset_per_share'
        },
        'sheet_names': ['ROE与净资产', 'ROE', '净资产']
    },
    'pe_valuation': {
        'required_fields': ['代码', '个股名称', '年份'],
        'optional_fields': ['PE最高值', 'PE中间值', 'PE最低值', '每股收益', '类型(actual/forecast)', '备注'],
        'model': PeValuation,
        'key_fields': ['company_id', 'year', 'type'],
        'field_mapping': {
            'PE最高值': 'pe_high',
            'PE中间值': 'pe_mid',
            'PE最低值': 'pe_low',
            '每股收益': 'eps',
            '类型(actual/forecast)': 'type',
            '备注': 'remark'
        },
        'sheet_names': ['PE估值', 'PE']
    },
    'shareholder_structure': {
        'required_fields': ['代码', '个股名称', '统计日期(YYYY-MM-DD)'],
        'optional_fields': ['股东类型', '持股比例(%)', '变动比例(%)'],
        'model': ShareholderStructure,
        'key_fields': ['company_id', 'stat_date', 'shareholder_type'],
        'field_mapping': {
            '统计日期(YYYY-MM-DD)': 'stat_date',
            '股东类型': 'shareholder_type',
            '持股比例(%)': 'holding_ratio',
            '变动比例(%)': 'change_ratio'
        },
        'sheet_names': ['股东结构']
    },
    'shareholder_count': {
        'required_fields': ['代码', '个股名称', '统计日期(YYYY-MM-DD)'],
        'optional_fields': ['股东总人数', '较上期变化'],
        'model': ShareholderCount,
        'key_fields': ['company_id', 'stat_date'],
        'field_mapping': {
            '统计日期(YYYY-MM-DD)': 'stat_date',
            '股东总人数': 'total_holders',
            '较上期变化': 'change'
        },
        'sheet_names': ['股东户数']
    },
    'rd_expense': {
        'required_fields': ['代码', '个股名称', '年份'],
        'optional_fields': ['主营收入(元)', '研发费用(元)', '研发费用占比(%)', '费用增长率', '费用回报率'],
        'model': RdExpense,
        'key_fields': ['company_id', 'year'],
        'field_mapping': {
            '主营收入(元)': 'revenue',
            '研发费用(元)': 'rd_expense',
            '研发费用占比(%)': 'rd_ratio',
            '费用增长率': 'rd_growth',
            '费用回报率': 'rd_return'
        },
        'sheet_names': ['研发投入', '研发']
    },
    'rd_staff': {
        'required_fields': ['代码', '个股名称', '年份'],
        'optional_fields': ['研发人员规模', '同比增长', '占员工总数(%)', '本科人数', '硕士人数', '本科+硕士占比(%)', '备注'],
        'model': RdStaff,
        'key_fields': ['company_id', 'year'],
        'field_mapping': {
            '研发人员规模': 'staff_count',
            '同比增长': 'growth',
            '占员工总数(%)': 'percent_of_total',
            '本科人数': 'bachelor',
            '硕士人数': 'master',
            '本科+硕士占比(%)': 'bachelor_master_ratio',
            '备注': 'remark'
        },
        'sheet_names': ['研发人员', '研发人员']
    }
}

# 工作表名称到模块key的映射
SHEET_NAME_TO_MODULE = {}
for module_key, schema in MODULE_SCHEMA.items():
    for sheet_name in schema.get('sheet_names', []):
        SHEET_NAME_TO_MODULE[sheet_name] = module_key


def parse_date(value):
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        try:
            return datetime.strptime(value, '%Y-%m-%d').date()
        except:
            return None
    return None


def parse_number(value):
    if value is None or value == '':
        return None
    if isinstance(value, str):
        if '%' in value:
            try:
                return float(value.replace('%', ''))
            except:
                return None
        try:
            return float(value)
        except:
            return None
    try:
        return float(value)
    except:
        return None


def get_or_create_company(code, name):
    company = Company.query.filter_by(code=code).first()
    if not company:
        company = Company(code=code, name=name)
        db.session.add(company)
        db.session.flush()
    return company


def read_excel_to_list(file_content):
    from openpyxl import load_workbook
    import tempfile
    import os
    
    temp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            file_content.seek(0)
            temp_file.write(file_content.read())
            temp_path = temp_file.name
        
        wb = load_workbook(temp_path, read_only=True, data_only=True)
        result = {}
        
        # 检查是否有工作表
        if not wb.sheetnames:
            wb.close()
            return result
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            data = []
            headers = []
            
            # 检查工作表是否有数据
            try:
                first_row = list(ws[1])
                if not first_row:
                    continue
                for cell in first_row:
                    headers.append(cell.value)
            except:
                # 空工作表或无法读取
                continue
            
            # 检查是否有表头
            if not headers or all(h is None for h in headers):
                continue
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                row_data = {}
                for col_idx, cell in enumerate(row):
                    if col_idx < len(headers):
                        row_data[headers[col_idx]] = cell.value
                data.append(row_data)
            
            result[sheet_name] = data
        
        wb.close()
        return result
    finally:
        if temp_path and os.path.exists(temp_path):
            try:
                os.unlink(temp_path)
            except:
                pass


def process_module_data(data_list, module_name, preview=False):
    schema = MODULE_SCHEMA.get(module_name)
    if not schema:
        return {'success': False, 'error': f'未知模块: {module_name}'}
    
    required_fields = schema['required_fields']
    field_mapping = schema['field_mapping']
    model = schema['model']
    key_fields = schema['key_fields']
    
    if data_list:
        columns = set(data_list[0].keys())
        missing_fields = [f for f in required_fields if f not in columns]
        if missing_fields:
            return {'success': False, 'error': f'缺少必需字段: {missing_fields}'}
    
    results = []
    errors = []
    success_count = 0
    
    for idx, row in enumerate(data_list):
        try:
            code = str(row.get('代码', '')).strip() if row.get('代码') else ''
            name = str(row.get('个股名称', '')).strip() if row.get('个股名称') else ''
            
            if not code or not name:
                errors.append(f'第{idx+2}行: 代码或个股名称为空')
                continue
            
            if not preview:
                company = get_or_create_company(code, name)
            else:
                class DummyObj:
                    id = 0
                company = DummyObj()
            
            data = {'company_id': company.id}
            
            for excel_field, db_field in field_mapping.items():
                if excel_field in row:
                    value = row.get(excel_field)
                    
                    if '日期' in excel_field:
                        value = parse_date(value)
                    elif db_field in ['type']:
                        value = str(value).strip() if value else 'actual'
                    elif db_field in ['remark', 'shareholder_type']:
                        value = str(value).strip() if value else ''
                    else:
                        value = parse_number(value)
                    
                    data[db_field] = value
            
            if '年份' in row:
                year_val = row.get('年份')
                if year_val:
                    try:
                        data['year'] = int(year_val)
                    except:
                        pass
            
            if preview:
                results.append({
                    'row': idx + 2,
                    'code': code,
                    'name': name,
                    'data': data
                })
                success_count += 1
                continue
            
            query = model.query
            for key in key_fields:
                if key in data:
                    query = query.filter(getattr(model, key) == data[key])
            
            existing = query.first()
            
            if existing:
                for key, value in data.items():
                    if key != 'company_id' and value is not None:
                        setattr(existing, key, value)
            else:
                new_record = model(**data)
                db.session.add(new_record)
            
            success_count += 1
            
        except Exception as e:
            errors.append(f'第{idx+2}行: {str(e)}')
    
    return {
        'success': True,
        'total': len(data_list),
        'success_count': success_count,
        'error_count': len(errors),
        'errors': errors,
        'preview_data': results if preview else None
    }


def merge_and_import_data(data_dict, preview=False):
    all_results = {}
    total_success = 0
    total_error = 0
    all_errors = []
    
    for module_name, data_list in data_dict.items():
        if not data_list:
            continue
        
        result = process_module_data(data_list, module_name, preview)
        
        if not result['success']:
            all_results[module_name] = result
            all_errors.append(f'[{module_name}] {result["error"]}')
            continue
        
        all_results[module_name] = result
        total_success += result['success_count']
        total_error += result['error_count']
        all_errors.extend([f'[{module_name}] {e}' for e in result['errors']])
    
    if not preview:
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return {
                'success': False,
                'error': f'数据库提交失败: {str(e)}',
                'details': all_results
            }
    
    return {
        'success': True,
        'total_success': total_success,
        'total_error': total_error,
        'errors': all_errors[:50],
        'details': all_results
    }


@api_bp.route('/data_import/upload', methods=['POST'])
def upload_and_preview():
    if 'file' not in request.files:
        return jsonify({'code': 400, 'message': '未找到文件'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'code': 400, 'message': '未选择文件'}), 400
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'code': 400, 'message': '仅支持Excel文件'}), 400
    
    single_module = request.form.get('module')
    
    try:
        import io
        file_content = io.BytesIO(file.read())
        all_sheets = read_excel_to_list(file_content)
        sheet_names = list(all_sheets.keys())
        
        # 检查是否有有效的工作表
        if not sheet_names:
            return jsonify({'code': 400, 'message': 'Excel文件为空或没有有效数据，请检查文件内容'}), 400
        
        data_dict = {}
        
        if len(sheet_names) > 1:
            # 多Sheet文件 - 使用工作表名称映射
            for sheet_name in sheet_names:
                module_key = SHEET_NAME_TO_MODULE.get(sheet_name)
                
                if module_key:
                    data_list = all_sheets[sheet_name]
                    if data_list:
                        data_dict[module_key] = data_list
        else:
            # 单Sheet文件
            sheet_name = sheet_names[0]
            data_list = all_sheets[sheet_name]
            
            # 检查工作表是否有数据
            if not data_list:
                return jsonify({'code': 400, 'message': f'工作表 "{sheet_name}" 中没有数据，请检查文件内容'}), 400
            
            if single_module and single_module in MODULE_SCHEMA:
                data_dict[single_module] = data_list
            else:
                # 尝试通过列名识别模块
                columns = set(data_list[0].keys()) if data_list else set()
                
                # 检查是否有列名
                if not columns or all(c is None for c in columns):
                    return jsonify({'code': 400, 'message': f'工作表 "{sheet_name}" 中没有有效的列名，请检查第一行是否包含表头'}), 400
                
                for module_name, schema in MODULE_SCHEMA.items():
                    required = set(schema['required_fields'])
                    if required.issubset(columns):
                        data_dict[module_name] = data_list
                        break
                
                if not data_dict:
                    # 提供更详细的错误信息
                    missing_fields_info = []
                    for module_name, schema in MODULE_SCHEMA.items():
                        required = set(schema['required_fields'])
                        missing = required - columns
                        if missing and len(missing) < len(required):  # 有部分匹配
                            missing_fields_info.append(f"缺少字段: {', '.join(missing)}")
                    
                    if missing_fields_info:
                        return jsonify({
                            'code': 400, 
                            'message': '数据格式不正确，请检查以下问题:\n' + '\n'.join(missing_fields_info[:3])
                        }), 400
                    
                    available_columns = ', '.join(str(c) for c in columns if c) if columns else '无'
                    return jsonify({
                        'code': 400, 
                        'message': f'无法识别数据模块。当前列: {available_columns}。请使用正确的工作表名称或指定模块类型。'
                    }), 400
        
        if not data_dict:
            # 检查是否有工作表
            if not sheet_names:
                return jsonify({'code': 400, 'message': 'Excel文件为空，请检查文件内容'}), 400
            
            # 检查工作表是否有数据
            has_data = any(len(all_sheets[sheet]) > 0 for sheet in sheet_names)
            if not has_data:
                return jsonify({'code': 400, 'message': 'Excel文件中没有数据，请检查文件内容'}), 400
            
            # 检查是否缺少必填字段
            missing_fields_info = []
            for sheet_name in sheet_names:
                data_list = all_sheets.get(sheet_name, [])
                if data_list:
                    columns = set(data_list[0].keys())
                    for module_name, schema in MODULE_SCHEMA.items():
                        required = set(schema['required_fields'])
                        missing = required - columns
                        if missing and len(missing) < len(required):  # 有部分匹配
                            missing_fields_info.append(f"{sheet_name}: 缺少字段 {', '.join(missing)}")
            
            if missing_fields_info:
                return jsonify({
                    'code': 400, 
                    'message': '数据格式不正确，请检查以下问题:\n' + '\n'.join(missing_fields_info)
                }), 400
            
            return jsonify({'code': 400, 'message': '未找到有效数据，请检查工作表名称或列名是否正确'}), 400
        
        result = merge_and_import_data(data_dict, preview=True)
        
        if not result['success']:
            return jsonify({'code': 500, 'message': result['error']}), 500
        
        return jsonify({
            'code': 200,
            'message': '预览成功',
            'data': {
                'modules': list(data_dict.keys()),
                'preview': result['details'],
                'summary': {
                    'total_modules': len(data_dict),
                    'total_records': sum(r['success_count'] for r in result['details'].values())
                }
            }
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'code': 500, 'message': f'处理失败: {str(e)}'}), 500


@api_bp.route('/data_import/import', methods=['POST'])
def execute_import():
    if 'file' not in request.files:
        return jsonify({'code': 400, 'message': '未找到文件'}), 400
    
    file = request.files['file']
    single_module = request.form.get('module')
    
    try:
        import io
        file_content = io.BytesIO(file.read())
        all_sheets = read_excel_to_list(file_content)
        sheet_names = list(all_sheets.keys())
        
        # 检查是否有有效的工作表
        if not sheet_names:
            return jsonify({'code': 400, 'message': 'Excel文件为空或没有有效数据，请检查文件内容'}), 400
        
        data_dict = {}
        
        if len(sheet_names) > 1:
            # 多Sheet文件 - 使用工作表名称映射
            for sheet_name in sheet_names:
                module_key = SHEET_NAME_TO_MODULE.get(sheet_name)
                
                if module_key:
                    data_list = all_sheets[sheet_name]
                    if data_list:
                        data_dict[module_key] = data_list
        else:
            # 单Sheet文件
            sheet_name = sheet_names[0]
            data_list = all_sheets[sheet_name]
            
            # 检查工作表是否有数据
            if not data_list:
                return jsonify({'code': 400, 'message': f'工作表 "{sheet_name}" 中没有数据，请检查文件内容'}), 400
            
            if single_module and single_module in MODULE_SCHEMA:
                data_dict[single_module] = data_list
            else:
                # 尝试通过列名识别模块
                columns = set(data_list[0].keys()) if data_list else set()
                
                # 检查是否有列名
                if not columns or all(c is None for c in columns):
                    return jsonify({'code': 400, 'message': f'工作表 "{sheet_name}" 中没有有效的列名，请检查第一行是否包含表头'}), 400
                
                for module_name, schema in MODULE_SCHEMA.items():
                    required = set(schema['required_fields'])
                    if required.issubset(columns):
                        data_dict[module_name] = data_list
                        break
                
                if not data_dict:
                    # 提供更详细的错误信息
                    missing_fields_info = []
                    for module_name, schema in MODULE_SCHEMA.items():
                        required = set(schema['required_fields'])
                        missing = required - columns
                        if missing and len(missing) < len(required):  # 有部分匹配
                            missing_fields_info.append(f"缺少字段: {', '.join(missing)}")
                    
                    if missing_fields_info:
                        return jsonify({
                            'code': 400, 
                            'message': '数据格式不正确，请检查以下问题:\n' + '\n'.join(missing_fields_info[:3])
                        }), 400
                    
                    available_columns = ', '.join(str(c) for c in columns if c) if columns else '无'
                    return jsonify({
                        'code': 400, 
                        'message': f'无法识别数据模块。当前列: {available_columns}。请使用正确的工作表名称或指定模块类型。'
                    }), 400
        
        if not data_dict:
            # 检查工作表是否有数据
            has_data = any(len(all_sheets[sheet]) > 0 for sheet in sheet_names)
            if not has_data:
                return jsonify({'code': 400, 'message': 'Excel文件中没有数据，请检查文件内容'}), 400
            
            # 检查是否缺少必填字段
            missing_fields_info = []
            for sheet_name in sheet_names:
                data_list = all_sheets.get(sheet_name, [])
                if data_list:
                    columns = set(data_list[0].keys())
                    for module_name, schema in MODULE_SCHEMA.items():
                        required = set(schema['required_fields'])
                        missing = required - columns
                        if missing and len(missing) < len(required):  # 有部分匹配
                            missing_fields_info.append(f"{sheet_name}: 缺少字段 {', '.join(missing)}")
            
            if missing_fields_info:
                return jsonify({
                    'code': 400, 
                    'message': '数据格式不正确，请检查以下问题:\n' + '\n'.join(missing_fields_info)
                }), 400
            
            return jsonify({'code': 400, 'message': '未找到有效数据，请检查工作表名称或列名是否正确'}), 400
        
        result = merge_and_import_data(data_dict, preview=False)
        
        if not result['success']:
            return jsonify({'code': 500, 'message': result['error']}), 500
        
        return jsonify({
            'code': 200,
            'message': '导入成功',
            'data': {
                'modules': list(data_dict.keys()),
                'total_success': result['total_success'],
                'total_error': result['total_error'],
                'errors': result['errors'],
                'details': {k: {'success': v['success_count'], 'error': v['error_count']} 
                           for k, v in result['details'].items()}
            }
        })
        
    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({'code': 500, 'message': f'导入失败: {str(e)}'}), 500


@api_bp.route('/data_import/template', methods=['GET'])
def download_merged_template():
    import io
    from flask import send_file
    from openpyxl import Workbook
    
    output = io.BytesIO()
    wb = Workbook()
    wb.remove(wb.active)
    
    sheet_names = {
        'profit_rate': '利润率',
        'non_recurring': '扣非净利润',
        'roe_net_asset': 'ROE与净资产',
        'pe_valuation': 'PE估值',
        'shareholder_structure': '股东结构',
        'shareholder_count': '股东户数',
        'rd_expense': '研发投入',
        'rd_staff': '研发人员'
    }
    
    for module_name, schema in MODULE_SCHEMA.items():
        columns = schema['required_fields'] + schema['optional_fields']
        sample_data = {}
        for col in columns:
            if '代码' in col:
                sample_data[col] = '600584'
            elif '名称' in col:
                sample_data[col] = '长电科技'
            elif '年份' in col:
                sample_data[col] = 2024
            elif '日期' in col:
                sample_data[col] = '2024-06-30'
            elif '类型' in col:
                sample_data[col] = 'actual'
            else:
                sample_data[col] = 0
        
        ws = wb.create_sheet(title=sheet_names.get(module_name, module_name))
        
        for c_idx, col_name in enumerate(columns, 1):
            ws.cell(row=1, column=c_idx, value=col_name)
        
        for c_idx, col_name in enumerate(columns, 1):
            ws.cell(row=2, column=c_idx, value=sample_data.get(col_name, ''))
    
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='数据导入模板_全模块.xlsx'
    )
